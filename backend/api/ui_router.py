"""API для отдельного UI-клиента ручной проверки флоу."""

from __future__ import annotations

import asyncio
import csv
import io
import json
import re
import uuid
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
import os

from api.task_dispatch import (
    _get_embedding_client,
    dispatch_clusterize,
    dispatch_memory_save,
    dispatch_normalize,
    fetch_task_status,
)
from infrastructure.naming.enriched_name import collapse_cluster_rows, render_template
from infrastructure.naming.merge_separators import normalize_merge_separator
from infrastructure.llm.factory import resolve_llm_provider
from pydantic import BaseModel, Field
from schemas.memory import MemoryItem, MemorySaveRequest
from schemas.task import (
    ClusterInput,
    ClusterProfileProvider,
    ClusterizeRequest,
    EmbeddingProvider,
    NormalizeProvider,
    NormalizeRequest,
)

router = APIRouter(tags=["ui"])


def _normalize_attribute_merge(raw: dict[str, Any] | None) -> dict[str, str]:
    return {
        str(key).strip(): str(value).strip()
        for key, value in dict(raw or {}).items()
        if str(key).strip() and str(value).strip() in ("priority", "accumulative")
    }


def _normalize_attribute_merge_separators(raw: dict[str, Any] | None) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in dict(raw or {}).items():
        attr = str(key).strip()
        if not attr:
            continue
        separator = normalize_merge_separator(str(value))
        if separator:
            normalized[attr] = separator
    return normalized


@dataclass
class SessionData:
    """Сессионное состояние UI-процесса."""

    base_url: str = "http://127.0.0.1:8000"
    headers: list[str] = field(default_factory=list)
    cleaned_rows: list[dict[str, str]] = field(default_factory=list)
    excluded_row_indices: set[int] = field(default_factory=set)
    selected_column: str | None = None
    base_attributes: list[str] = field(default_factory=list)
    items: list[str] = field(default_factory=list)
    embedding_provider: str = "local"
    cluster_profile_provider: str = "g4f"
    normalize_provider: str = "g4f"
    clusterize_task_id: str | None = None
    clusterize_result: dict[str, Any] | None = None
    approved_clusters: list[dict[str, Any]] = field(default_factory=list)
    normalize_task_id: str | None = None
    normalize_result: dict[str, Any] | None = None


SESSIONS: dict[str, SessionData] = {}


class StartTaskPayload(BaseModel):
    session_id: str
    base_url: str = Field(default="http://127.0.0.1:8000")
    provider: str | None = None
    mode: str = Field(default="start", description="start|resume|restart")
    cluster_indexes: list[int] = Field(default_factory=list)
    cluster_attribute_mode: str = Field(default="default", description="default|all|missing")


class ClusterizeStartPayload(BaseModel):
    session_id: str
    base_url: str = Field(default="http://127.0.0.1:8000")
    embedding_provider: str = "local"
    profile_provider: str = "g4f"


class ConfigurePayload(BaseModel):
    session_id: str
    selected_column: str
    base_attributes: list[str]


class ClustersPayload(BaseModel):
    session_id: str
    clusters: list[dict[str, Any]]


class SaveMemoryPayload(BaseModel):
    session_id: str
    base_url: str = Field(default="http://127.0.0.1:8000")
    cluster_index: int | None = None


class SessionPayload(BaseModel):
    session_id: str


class EnsureSessionPayload(BaseModel):
    session_id: str | None = None


class RowIncludePayload(BaseModel):
    session_id: str
    row_index: int
    included: bool


class RowUpdatePayload(BaseModel):
    session_id: str
    row_index: int
    cells: dict[str, str] = Field(default_factory=dict)


class RowCreatePayload(BaseModel):
    session_id: str
    cells: dict[str, str] = Field(default_factory=dict)


class RowDeletePayload(BaseModel):
    session_id: str
    row_index: int


class ColumnAddPayload(BaseModel):
    session_id: str
    column_name: str
    after_column: str | None = None


class SetSourceColumnPayload(BaseModel):
    session_id: str
    column_name: str


class ColumnRenamePayload(BaseModel):
    session_id: str
    old_name: str
    new_name: str


class ColumnDeletePayload(BaseModel):
    session_id: str
    column_name: str


class MemoryClusterLoadPayload(BaseModel):
    cluster_name: str


class MemoryClusterDeletePayload(BaseModel):
    cluster_name: str


class MemoryItemDeletePayload(BaseModel):
    text: str


class MemorySearchPayload(BaseModel):
    query: str
    limit: int = Field(default=20, ge=1, le=100)


def _get_session(session_id: str) -> SessionData:
    session = SESSIONS.get(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")
    return session


def _reset_session_state(session: SessionData) -> None:
    """Сбрасывает прогон, оставляя сессию и базовый URL."""
    session.headers = []
    session.cleaned_rows = []
    session.excluded_row_indices = set()
    session.selected_column = None
    session.base_attributes = []
    session.items = []
    session.clusterize_task_id = None
    session.clusterize_result = None
    session.approved_clusters = []
    session.normalize_task_id = None
    session.normalize_result = None


def _row_empty(row: dict[str, str]) -> bool:
    return all(not str(v).strip() for v in row.values())


def _validate_llm_provider(provider: str) -> str:
    resolved = resolve_llm_provider(provider)
    if resolved == "gemini" and not os.getenv("GEMINI_API_KEY"):
        raise HTTPException(
            status_code=400,
            detail="GEMINI_API_KEY не задан на сервере. Укажите ключ в .env или выберите g4f.",
        )
    return resolved


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], []
    headers = [str(h).strip() for h in reader.fieldnames]
    rows: list[dict[str, str]] = []
    for row in reader:
        rows.append({str(k).strip(): str(v or "").strip() for k, v in row.items()})
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(v).strip() if v is not None else f"col_{i+1}" for i, v in enumerate(rows[0])]
    out_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        out_rows.append({headers[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(row)})
    return headers, out_rows


def _clean_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], int, int]:
    cleaned: list[dict[str, str]] = []
    seen: set[tuple[tuple[str, str], ...]] = set()
    removed_empty = 0
    removed_duplicates = 0
    for row in rows:
        if _row_empty(row):
            removed_empty += 1
            continue
        key = tuple(sorted((k, str(v).strip()) for k, v in row.items()))
        if key in seen:
            removed_duplicates += 1
            continue
        seen.add(key)
        cleaned.append({k: str(v).strip() for k, v in row.items()})
    return cleaned, removed_empty, removed_duplicates


def _item_cluster_name_map(clusters: list[dict[str, Any]]) -> dict[str, str]:
    """Строит соответствие номенклатура/обогащённое имя → имя кластера."""
    mapping: dict[str, str] = {}
    for cluster in clusters:
        name = str(cluster.get("name", "")).strip() or "Cluster"
        item_names = [str(i).strip() for i in cluster.get("items", []) if str(i).strip()]
        for row in cluster.get("rows") or []:
            enriched = str(row.get("enriched_name") or "").strip()
            if enriched:
                mapping[enriched] = name
            for alias in row.get("aliases") or []:
                alias_name = str(alias).strip()
                if alias_name:
                    mapping[alias_name] = name
            item_name = str(row.get("item", "")).strip()
            if item_name:
                item_names.append(item_name)
        for item_name in item_names:
            mapping[item_name] = name
    return mapping


def _row_entries_for_rededupe(row: dict[str, Any], template: str) -> list[dict[str, Any]]:
    """Разворачивает строку кластера в позиции для повторной дедупликации."""
    entries: list[dict[str, Any]] = []
    row_source = str(row.get("source") or "ai").strip().lower()
    if row_source not in ("memory", "ai"):
        row_source = "ai"
    row_values = {
        str(key).strip(): str(value).strip()
        for key, value in dict(row.get("values") or {}).items()
        if str(key).strip()
    }
    enriched_row = str(row.get("enriched_name") or "").strip()
    members = row.get("members") or []
    if members:
        for member in members:
            item_name = str(member.get("item", "")).strip()
            if not item_name:
                continue
            values = {
                str(key).strip(): str(value).strip()
                for key, value in dict(member.get("values") or row_values).items()
                if str(key).strip()
            }
            source = str(member.get("source") or row_source).strip().lower()
            if source not in ("memory", "ai"):
                source = row_source
            enriched = render_template(template, values) if template else enriched_row or item_name
            entries.append(
                {
                    "enriched_name": enriched,
                    "item": item_name,
                    "values": values,
                    "source": source,
                }
            )
        return entries

    aliases = [str(alias).strip() for alias in row.get("aliases") or [] if str(alias).strip()]
    if not aliases:
        item_name = str(row.get("item", "")).strip()
        if item_name:
            aliases = [item_name]
    for alias in aliases:
        enriched = render_template(template, row_values) if template else enriched_row or alias
        entries.append(
            {
                "enriched_name": enriched,
                "item": alias,
                "values": dict(row_values),
                "source": row_source,
            }
        )
    return entries


def _apply_collapsed_clusters(session: SessionData) -> None:
    """Подменяет approved_clusters схлопнутыми строками после нормализации."""
    collapsed = (session.normalize_result or {}).get("clusters_collapsed") or []
    if not collapsed:
        return
    by_name = {str(c.get("name", "")).strip(): c for c in collapsed if str(c.get("name", "")).strip()}
    updated: list[dict[str, Any]] = []
    for cluster in session.approved_clusters or []:
        name = str(cluster.get("name", "")).strip()
        if name in by_name:
            merged = dict(cluster)
            merged.update(by_name[name])
            updated.append(merged)
        else:
            updated.append(cluster)
    session.approved_clusters = updated


def _normalize_mode(raw_mode: str | None) -> str:
    mode = str(raw_mode or "start").strip().lower()
    if mode not in {"start", "resume", "restart"}:
        raise HTTPException(status_code=400, detail="mode must be start|resume|restart")
    return mode


def _build_normalize_clusters_payload(session: SessionData) -> list[ClusterInput]:
    """Преобразует approved_clusters в payload для NormalizeRequest."""
    clusters: list[ClusterInput] = []
    for cluster in session.approved_clusters:
        item_sources: dict[str, str] = {}
        item_values: dict[str, dict[str, str]] = {}
        items: list[str] = []
        seen_items: set[str] = set()
        for row in cluster.get("rows") or []:
            row_source = str(row.get("source") or "ai").strip().lower()
            source = row_source if row_source in ("memory", "ai") else "ai"
            row_values = {
                str(key).strip(): str(value).strip()
                for key, value in dict(row.get("values") or {}).items()
                if str(key).strip()
            }
            members = row.get("members") or []
            if members:
                for member in members:
                    item_name = str(member.get("item", "")).strip()
                    if not item_name:
                        continue
                    if item_name not in seen_items:
                        seen_items.add(item_name)
                        items.append(item_name)
                    member_source_raw = str(member.get("source") or source).strip().lower()
                    item_sources[item_name] = (
                        member_source_raw if member_source_raw in ("memory", "ai") else source
                    )
                    item_values[item_name] = {
                        str(key).strip(): str(value).strip()
                        for key, value in dict(member.get("values") or row_values).items()
                        if str(key).strip()
                    }
                continue

            item_name = str(row.get("item", "")).strip()
            if not item_name:
                continue
            if item_name not in seen_items:
                seen_items.add(item_name)
                items.append(item_name)
            item_sources[item_name] = source
            item_values[item_name] = dict(row_values)
        if not items:
            items = [str(i).strip() for i in cluster.get("items", []) if str(i).strip()]
            for item_name in items:
                item_sources[item_name] = "ai"
                item_values[item_name] = {}
        attribute_merge = _normalize_attribute_merge(cluster.get("attribute_merge"))
        attribute_merge_separators = _normalize_attribute_merge_separators(
            cluster.get("attribute_merge_separators")
        )
        clusters.append(
            ClusterInput(
                name=str(cluster.get("name", "")).strip() or "Cluster",
                attributes=[str(a).strip() for a in cluster.get("attributes", []) if str(a).strip()],
                items=items,
                enriched_name_template=str(cluster.get("enriched_name_template") or "").strip(),
                item_sources=item_sources,
                item_values=item_values,
                attribute_merge=attribute_merge,
                attribute_merge_separators=attribute_merge_separators,
            )
        )
    return clusters


def _extract_resume_indexes(result: dict[str, Any], total_clusters: int) -> list[int]:
    """Возвращает валидированные индексы обработанных кластеров."""
    indexes: list[int] = []
    for index in list(result.get("completed_cluster_indexes") or []):
        try:
            value = int(index)
        except (TypeError, ValueError):
            continue
        if 0 <= value < total_clusters and value not in indexes:
            indexes.append(value)
    return sorted(indexes)


def _normalize_cluster_attribute_mode(raw_mode: str | None) -> str:
    mode = str(raw_mode or "default").strip().lower()
    if mode not in {"default", "all", "missing"}:
        raise HTTPException(status_code=400, detail="cluster_attribute_mode must be default|all|missing")
    return mode


def _normalize_selected_cluster_indexes(raw_indexes: list[int], total_clusters: int) -> list[int]:
    if not raw_indexes:
        return list(range(total_clusters))
    indexes: list[int] = []
    for raw in raw_indexes:
        try:
            value = int(raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="cluster_indexes must contain integers") from None
        if value < 0 or value >= total_clusters:
            raise HTTPException(status_code=400, detail=f"cluster index out of range: {value}")
        if value not in indexes:
            indexes.append(value)
    return indexes


def _clusters_from_known_items(
    known_items: list[dict[str, Any]],
    base_attributes: list[str],
) -> list[dict[str, Any]]:
    """Группирует позиции из памяти по cluster_name в enriched-формате."""
    from collections import defaultdict

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for known in known_items:
        item_name = str(known.get("item", "")).strip()
        if not item_name:
            continue
        cluster_name = str(known.get("cluster_name") or "").strip()
        if not cluster_name:
            cluster_name = str(known.get("text") or "").strip() or item_name
        groups[cluster_name].append(known)

    clusters: list[dict[str, Any]] = []
    for cluster_name, group in groups.items():
        attr_keys: list[str] = []
        rows_by_enriched: dict[str, dict[str, Any]] = {}
        cluster_attribute_merge: dict[str, str] = {}
        cluster_attribute_merge_separators: dict[str, str] = {}
        for known in group:
            item_name = str(known.get("item", "")).strip()
            if not item_name:
                continue
            attrs = known.get("attributes")
            values: dict[str, str] = {}
            if isinstance(attrs, dict):
                values = {
                    str(key).strip(): str(value).strip()
                    for key, value in attrs.items()
                    if str(key).strip()
                }
            known_attribute_merge = _normalize_attribute_merge(known.get("attribute_merge"))
            for key, value in known_attribute_merge.items():
                if key not in cluster_attribute_merge:
                    cluster_attribute_merge[key] = value
            known_attribute_merge_separators = _normalize_attribute_merge_separators(
                known.get("attribute_merge_separators")
            )
            for key, value in known_attribute_merge_separators.items():
                if key not in cluster_attribute_merge_separators:
                    cluster_attribute_merge_separators[key] = value
            for key in values:
                if key not in attr_keys:
                    attr_keys.append(key)
            original_item_values = {
                str(alias).strip(): {
                    str(key).strip(): str(value).strip()
                    for key, value in dict(member_values or {}).items()
                    if str(key).strip()
                }
                for alias, member_values in dict(known.get("original_item_values") or {}).items()
                if str(alias).strip()
            }
            enriched_name = (
                str(known.get("enriched_name") or known.get("text") or "").strip() or item_name
            )
            originals = [
                str(alias).strip()
                for alias in known.get("original_items") or []
                if str(alias).strip()
            ]
            aliases = originals or [item_name]
            row = rows_by_enriched.get(enriched_name)
            if row is None:
                row = {
                    "enriched_name": enriched_name,
                    "aliases": [],
                    "item": aliases[0],
                    "values": dict(values),
                    "source": "memory",
                    "members": [],
                }
                rows_by_enriched[enriched_name] = row
            else:
                merged_values = dict(row.get("values") or {})
                for key, value in values.items():
                    if key not in merged_values or not str(merged_values[key]).strip():
                        merged_values[key] = value
                row["values"] = merged_values

            existing_aliases = list(row.get("aliases") or [])
            for alias in aliases:
                if alias not in existing_aliases:
                    existing_aliases.append(alias)
            row["aliases"] = existing_aliases
            members = list(row.get("members") or [])
            member_by_item = {str(member.get("item", "")).strip(): member for member in members}
            for alias in aliases:
                alias_values = dict(original_item_values.get(alias) or values)
                for key in alias_values:
                    if key not in attr_keys:
                        attr_keys.append(key)
                existing_member = member_by_item.get(alias)
                if alias and existing_member is None:
                    members.append(
                        {
                            "item": alias,
                            "values": alias_values,
                            "source": "memory",
                        }
                    )
                    member_by_item[alias] = members[-1]
                elif alias and existing_member is not None:
                    merged_member_values = dict(existing_member.get("values") or {})
                    for key, value in alias_values.items():
                        if key not in merged_member_values or not str(merged_member_values[key]).strip():
                            merged_member_values[key] = value
                    existing_member["values"] = merged_member_values
            row["members"] = members

        rows = list(rows_by_enriched.values())
        items = [str(row.get("enriched_name") or "").strip() for row in rows if row.get("enriched_name")]
        template = ""
        for known in group:
            candidate = str(known.get("name_template") or "").strip()
            if candidate:
                template = candidate
                break
        clusters.append(
            {
                "name": cluster_name,
                "attributes": attr_keys or list(base_attributes),
                "enriched_name_template": template,
                "items": items,
                "rows": rows,
                "attribute_merge": cluster_attribute_merge,
                "attribute_merge_separators": cluster_attribute_merge_separators,
                "source": "memory",
                "memory_cluster_name": cluster_name,
            }
        )
    return clusters


def _build_default_clusters(clusterize_result: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not clusterize_result:
        return []
    base_attributes = list(clusterize_result.get("base_attributes") or [])
    clusters: list[dict[str, Any]] = []
    for idx, cluster in enumerate(clusterize_result.get("new_item_clusters", []), start=1):
        cluster_items = [
            str(item) for item in cluster.get("cluster_items", []) if str(item).strip()
        ]
        clusters.append(
            {
                "name": str(cluster.get("category") or f"Cluster {idx}"),
                "attributes": list(cluster.get("attributes") or base_attributes),
                "enriched_name_template": str(cluster.get("name_template") or "").strip(),
                "items": cluster_items,
                "rows": [
                    {"item": item_name, "values": {}, "source": "ai"}
                    for item_name in cluster_items
                ],
                "source": "ai",
                "memory_cluster_name": "",
            }
        )
    clusters.extend(
        _clusters_from_known_items(
            list(clusterize_result.get("known_items") or []),
            base_attributes,
        )
    )
    return clusters


def _memory_item_names(session: SessionData) -> set[str]:
    """Номенклатура, найденная в Qdrant при последней кластеризации."""
    names: set[str] = set()
    for known in (session.clusterize_result or {}).get("known_items") or []:
        item_name = str(known.get("item", "")).strip()
        if item_name:
            names.add(item_name)
        for alias in known.get("original_items") or []:
            alias_name = str(alias).strip()
            if alias_name:
                names.add(alias_name)
    return names


def _clusters_for_ui(session: SessionData) -> list[dict[str, Any]]:
    """Кластеры для UI: items + rows с values после нормализации."""
    clusters = session.approved_clusters or []
    memory_items = _memory_item_names(session)
    normalized_map: dict[str, dict[str, Any]] = {
        str(row.get("item", "")).strip(): dict(row.get("values") or {})
        for row in (session.normalize_result or {}).get("normalized") or []
        if str(row.get("item", "")).strip()
    }

    has_enriched_rows = any(
        str(row.get("enriched_name") or "").strip()
        for cluster in clusters
        for row in cluster.get("rows") or []
    )
    if has_enriched_rows:
        payload: list[dict[str, Any]] = []
        for cluster in clusters:
            name = str(cluster.get("name", "")).strip() or "Cluster"
            attrs = [str(a).strip() for a in cluster.get("attributes", []) if str(a).strip()]
            rows: list[dict[str, Any]] = []
            for row in cluster.get("rows") or []:
                enriched = str(row.get("enriched_name") or "").strip()
                if not enriched:
                    continue
                aliases = [
                    str(alias).strip()
                    for alias in row.get("aliases") or []
                    if str(alias).strip()
                ]
                if not aliases:
                    aliases = [str(row.get("item", "")).strip()] if row.get("item") else []
                values_raw = dict(row.get("values") or {})
                for alias in aliases:
                    values_raw = {**values_raw, **normalized_map.get(alias, {})}
                row_attrs = list(attrs)
                for key in values_raw:
                    key_name = str(key).strip()
                    if key_name and key_name not in row_attrs:
                        row_attrs.append(key_name)
                source = str(row.get("source") or "ai").strip().lower()
                if source not in ("memory", "ai"):
                    source = "ai"
                members_raw = row.get("members") or []
                members = [
                    {
                        "item": str(member.get("item", "")).strip(),
                        "values": {
                            str(k).strip(): str(v).strip()
                            for k, v in dict(member.get("values") or {}).items()
                            if str(k).strip()
                        },
                        "source": str(member.get("source") or source).strip().lower()
                        if str(member.get("source") or source).strip().lower()
                        in ("memory", "ai")
                        else source,
                    }
                    for member in members_raw
                    if str(member.get("item", "")).strip()
                ]
                rows.append(
                    {
                        "enriched_name": enriched,
                        "aliases": aliases or [enriched],
                        "item": aliases[0] if aliases else enriched,
                        "values": {
                            attr: str(values_raw.get(attr, "")).strip() for attr in row_attrs
                        },
                        "source": source,
                        "members": members,
                    }
                )
            merged_attrs = list(attrs)
            for row in rows:
                for attr in row["values"]:
                    if attr not in merged_attrs:
                        merged_attrs.append(attr)
            attribute_merge = _normalize_attribute_merge(cluster.get("attribute_merge"))
            attribute_merge_separators = _normalize_attribute_merge_separators(
                cluster.get("attribute_merge_separators")
            )
            payload.append(
                {
                    "name": name,
                    "attributes": merged_attrs or attrs,
                    "enriched_name_template": str(
                        cluster.get("enriched_name_template") or ""
                    ).strip(),
                    "attribute_merge": attribute_merge,
                    "attribute_merge_separators": attribute_merge_separators,
                    "items": [row["enriched_name"] for row in rows],
                    "rows": rows,
                    "source": _cluster_source(cluster),
                    "memory_cluster_name": str(cluster.get("memory_cluster_name") or "").strip(),
                }
            )
        return payload

    payload: list[dict[str, Any]] = []
    for cluster in clusters:
        name = str(cluster.get("name", "")).strip() or "Cluster"
        attrs = [str(a).strip() for a in cluster.get("attributes", []) if str(a).strip()]
        items = [str(i).strip() for i in cluster.get("items", []) if str(i).strip()]
        stored_rows_map: dict[str, dict[str, Any]] = {
            str(row.get("item", "")).strip(): row
            for row in cluster.get("rows") or []
            if str(row.get("item", "")).strip()
        }
        rows: list[dict[str, Any]] = []
        for item_name in items:
            stored_row = stored_rows_map.get(item_name, {})
            values_raw = {
                **dict(stored_row.get("values") or {}),
                **normalized_map.get(item_name, {}),
            }
            source = str(stored_row.get("source") or "").strip().lower()
            if source not in ("memory", "ai"):
                source = "memory" if item_name in memory_items else "ai"
            row_attrs = list(attrs)
            for key in values_raw:
                key_name = str(key).strip()
                if key_name and key_name not in row_attrs:
                    row_attrs.append(key_name)
            rows.append(
                {
                    "item": item_name,
                    "values": {attr: str(values_raw.get(attr, "")).strip() for attr in row_attrs},
                    "source": source,
                }
            )
        merged_attrs = list(attrs)
        for row in rows:
            for attr in row["values"]:
                if attr not in merged_attrs:
                    merged_attrs.append(attr)
        payload.append(
            {
                "name": name,
                "attributes": merged_attrs or attrs,
                "enriched_name_template": str(
                    cluster.get("enriched_name_template") or ""
                ).strip(),
                "items": items,
                "rows": rows,
                "source": _cluster_source(cluster),
                "memory_cluster_name": str(cluster.get("memory_cluster_name") or "").strip(),
            }
        )
    return payload


def _cluster_source(cluster: dict[str, Any]) -> str:
    source = str(cluster.get("source") or "").strip().lower()
    if source in ("memory", "ai", "manual"):
        return source
    row_sources = {
        str(row.get("source") or "").strip().lower()
        for row in cluster.get("rows") or []
        if str(row.get("source") or "").strip().lower() in ("memory", "ai")
    }
    if row_sources == {"memory"}:
        return "memory"
    if row_sources == {"ai"}:
        return "ai"
    if row_sources:
        return "manual"
    return "ai"


def _cluster_from_memory_points(
    cluster_name: str,
    points: list[dict[str, Any]],
    base_attributes: list[str],
) -> dict[str, Any]:
    cluster_attribute_merge: dict[str, str] = {}
    cluster_attribute_merge_separators: dict[str, str] = {}
    attr_keys: list[str] = []
    rows_by_enriched: dict[str, dict[str, Any]] = {}
    for point in points:
        enriched = str(point.get("text") or "").strip()
        if not enriched:
            continue
        if not cluster_attribute_merge:
            cluster_attribute_merge = _normalize_attribute_merge(point.get("attribute_merge"))
        if not cluster_attribute_merge_separators:
            cluster_attribute_merge_separators = _normalize_attribute_merge_separators(
                point.get("attribute_merge_separators")
            )
        originals = [
            str(alias).strip()
            for alias in point.get("original_items") or []
            if str(alias).strip()
        ]
        if not originals:
            originals = [enriched]
        values = {
            str(key).strip(): str(value).strip()
            for key, value in dict(point.get("attributes") or {}).items()
            if str(key).strip()
        }
        for key in values:
            if key not in attr_keys:
                attr_keys.append(key)
        original_item_values = {
            str(alias).strip(): {
                str(key).strip(): str(value).strip()
                for key, value in dict(item_values or {}).items()
                if str(key).strip()
            }
            for alias, item_values in dict(point.get("original_item_values") or {}).items()
            if str(alias).strip()
        }
        row = rows_by_enriched.get(enriched)
        if row is None:
            row = {
                "enriched_name": enriched,
                "aliases": [],
                "item": originals[0],
                "values": dict(values),
                "source": "memory",
                "members": [],
            }
            rows_by_enriched[enriched] = row
        existing_aliases = list(row.get("aliases") or [])
        existing_members = list(row.get("members") or [])
        existing_member_by_item = {
            str(member.get("item", "")).strip(): member
            for member in existing_members
            if str(member.get("item", "")).strip()
        }
        for alias in originals:
            if alias not in existing_aliases:
                existing_aliases.append(alias)
            member_values = dict(original_item_values.get(alias) or values)
            for key in member_values:
                if key not in attr_keys:
                    attr_keys.append(key)
            current_member = existing_member_by_item.get(alias)
            if current_member is None:
                existing_members.append(
                    {
                        "item": alias,
                        "values": member_values,
                        "source": "memory",
                    }
                )
                existing_member_by_item[alias] = existing_members[-1]
            else:
                merged_member_values = dict(current_member.get("values") or {})
                for key, value in member_values.items():
                    if key not in merged_member_values or not str(merged_member_values[key]).strip():
                        merged_member_values[key] = value
                current_member["values"] = merged_member_values
        row["aliases"] = existing_aliases
        row["members"] = existing_members
    rows = list(rows_by_enriched.values())
    if rows:
        return {
            "name": cluster_name,
            "attributes": attr_keys or list(base_attributes),
            "enriched_name_template": "",
            "items": [str(row.get("enriched_name") or "").strip() for row in rows if row.get("enriched_name")],
            "rows": rows,
            "attribute_merge": cluster_attribute_merge,
            "attribute_merge_separators": cluster_attribute_merge_separators,
            "source": "memory",
            "memory_cluster_name": cluster_name,
        }
    return {
        "name": cluster_name,
        "attributes": list(base_attributes),
        "enriched_name_template": "",
        "items": [],
        "rows": [],
        "attribute_merge": cluster_attribute_merge,
        "attribute_merge_separators": cluster_attribute_merge_separators,
        "source": "memory",
        "memory_cluster_name": cluster_name,
    }


def _normalized_map(session: SessionData) -> dict[str, dict[str, Any]]:
    normalized = (session.normalize_result or {}).get("normalized") or []
    return {
        str(row.get("item", "")).strip(): dict(row.get("values", {}))
        for row in normalized
        if str(row.get("item", "")).strip()
    }


def _cluster_export_rows(
    cluster: dict[str, Any],
    normalized_map: dict[str, dict[str, Any]],
) -> tuple[list[str], list[list[str]]]:
    attributes = [str(a).strip() for a in cluster.get("attributes", []) if str(a).strip()]
    rows_out: list[list[str]] = []
    items = [str(i).strip() for i in cluster.get("items", []) if str(i).strip()]
    for row in cluster.get("rows") or []:
        enriched = str(row.get("enriched_name") or "").strip()
        aliases = row.get("aliases") or []
        if enriched:
            values = dict(row.get("values") or {})
            for alias in aliases:
                values = {**values, **normalized_map.get(str(alias).strip(), {})}
            originals = "; ".join(str(a).strip() for a in aliases if str(a).strip())
            rows_out.append(
                [
                    enriched,
                    originals,
                    *[str(values.get(attr, "")).strip() for attr in attributes],
                ]
            )
            continue
    if rows_out:
        return attributes, rows_out
    for item_name in items:
        values = normalized_map.get(item_name, {})
        rows_out.append(
            [item_name, item_name, *[str(values.get(attr, "")).strip() for attr in attributes]]
        )
    return attributes, rows_out


def _get_cluster_for_export(session: SessionData, cluster_index: int) -> dict[str, Any]:
    clusters = session.approved_clusters or []
    if not clusters:
        raise HTTPException(status_code=400, detail="No clusters to export")
    if cluster_index < 0 or cluster_index >= len(clusters):
        raise HTTPException(status_code=400, detail="Cluster index out of range")
    return clusters[cluster_index]


def _safe_filename(name: str, fallback: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", str(name).strip())
    return cleaned or fallback


@router.post("/ui/api/session/new")
async def create_session() -> dict[str, str]:
    session_id = str(uuid.uuid4())
    SESSIONS[session_id] = SessionData()
    return {"session_id": session_id}


@router.post("/ui/api/session/ensure")
async def ensure_session(payload: EnsureSessionPayload) -> dict[str, str]:
    """Возвращает существующую сессию или создает новую."""
    session_id = (payload.session_id or "").strip()
    if session_id and session_id in SESSIONS:
        return {"session_id": session_id}
    new_session_id = str(uuid.uuid4())
    SESSIONS[new_session_id] = SessionData()
    return {"session_id": new_session_id}


@router.post("/ui/api/session/reset")
async def reset_session(payload: SessionPayload) -> dict[str, str]:
    session = _get_session(payload.session_id)
    _reset_session_state(session)
    return {"status": "reset"}


@router.post("/ui/api/session/drop")
async def drop_session(payload: SessionPayload) -> dict[str, str]:
    """Удаляет сессию полностью."""
    SESSIONS.pop(payload.session_id, None)
    return {"status": "dropped"}


@router.post("/ui/api/upload")
async def upload_data(
    session_id: str = Form(...),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    session = _get_session(session_id)
    ext = Path(file.filename or "").suffix.lower()
    content = await file.read()
    if ext == ".csv":
        headers, rows = _parse_csv(content)
    elif ext in {".xlsx", ".xlsm"}:
        headers, rows = _parse_xlsx(content)
    else:
        raise HTTPException(status_code=400, detail="Only CSV/XLSX supported")
    if not headers:
        raise HTTPException(status_code=400, detail="No columns found in file")
    cleaned_rows, removed_empty, removed_duplicates = _clean_rows(rows)
    if not cleaned_rows:
        raise HTTPException(status_code=400, detail="No rows left after cleanup")

    _reset_session_state(session)
    session.headers = headers
    session.cleaned_rows = cleaned_rows
    session.selected_column = headers[0] if headers else None

    return {
        "headers": headers,
        "stats": {
            "rows_total": len(rows),
            "rows_after_cleaning": len(cleaned_rows),
            "removed_empty_rows": removed_empty,
            "removed_duplicates": removed_duplicates,
        },
        "preview_rows": [],
    }


@router.get("/ui/api/rows/{session_id}")
async def list_rows(
    session_id: str,
    page: int = 1,
    page_size: int = 10,
) -> dict[str, Any]:
    session = _get_session(session_id)
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 10
    if page_size > 100:
        page_size = 100
    total = len(session.cleaned_rows)
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    end = start + page_size
    rows_payload: list[dict[str, Any]] = []
    for idx in range(start, min(end, total)):
        rows_payload.append(
            {
                "row_index": idx,
                "included": idx not in session.excluded_row_indices,
                "cells": session.cleaned_rows[idx],
            }
        )
    return {
        "page": page,
        "page_size": page_size,
        "total_rows": total,
        "total_pages": total_pages,
        "rows": rows_payload,
    }


@router.post("/ui/api/rows/include")
async def set_row_included(payload: RowIncludePayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if payload.row_index < 0 or payload.row_index >= len(session.cleaned_rows):
        raise HTTPException(status_code=400, detail="row_index out of range")
    if payload.included:
        session.excluded_row_indices.discard(payload.row_index)
    else:
        session.excluded_row_indices.add(payload.row_index)
    return {"status": "ok"}


@router.post("/ui/api/rows/update")
async def update_row(payload: RowUpdatePayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if payload.row_index < 0 or payload.row_index >= len(session.cleaned_rows):
        raise HTTPException(status_code=400, detail="row_index out of range")
    row = session.cleaned_rows[payload.row_index]
    for key, value in payload.cells.items():
        key_name = str(key).strip()
        if key_name in session.headers:
            row[key_name] = str(value or "").strip()
    return {"status": "ok", "row_index": payload.row_index, "cells": row}


@router.post("/ui/api/rows/add")
async def add_row(payload: RowCreatePayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if not session.headers:
        raise HTTPException(status_code=400, detail="Upload file first")
    new_row = {header: "" for header in session.headers}
    for key, value in payload.cells.items():
        key_name = str(key).strip()
        if key_name in session.headers:
            new_row[key_name] = str(value or "").strip()
    source_col = session.selected_column or session.headers[0]
    if not str(new_row.get(source_col, "")).strip():
        raise HTTPException(status_code=400, detail="Source nomenclature is required")
    session.cleaned_rows.insert(0, new_row)
    session.excluded_row_indices = {idx + 1 for idx in session.excluded_row_indices}
    return {"status": "ok", "row_index": 0, "cells": new_row}


@router.post("/ui/api/rows/delete")
async def delete_row(payload: RowDeletePayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if payload.row_index < 0 or payload.row_index >= len(session.cleaned_rows):
        raise HTTPException(status_code=400, detail="row_index out of range")
    session.cleaned_rows.pop(payload.row_index)
    updated_excluded: set[int] = set()
    for idx in session.excluded_row_indices:
        if idx == payload.row_index:
            continue
        updated_excluded.add(idx - 1 if idx > payload.row_index else idx)
    session.excluded_row_indices = updated_excluded
    return {"status": "ok"}


@router.post("/ui/api/columns/add")
async def add_column(payload: ColumnAddPayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    name = payload.column_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="column_name is required")
    if name in session.headers:
        raise HTTPException(status_code=400, detail="Column already exists")
    anchor = (payload.after_column or "").strip() or session.selected_column or ""
    if not anchor and session.headers:
        anchor = session.headers[0]
    if anchor and anchor in session.headers:
        insert_at = session.headers.index(anchor) + 1
    else:
        insert_at = 0
    session.headers.insert(insert_at, name)
    for row in session.cleaned_rows:
        row[name] = ""
    return {"status": "ok", "headers": session.headers}


@router.post("/ui/api/columns/set-source")
async def set_source_column(payload: SetSourceColumnPayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    name = payload.column_name.strip()
    if name not in session.headers:
        raise HTTPException(status_code=400, detail="Column not found")
    session.selected_column = name
    return {"selected_column": name, "headers": session.headers}


@router.post("/ui/api/columns/rename")
async def rename_column(payload: ColumnRenamePayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    old_name = payload.old_name.strip()
    new_name = payload.new_name.strip()
    if not old_name or not new_name:
        raise HTTPException(status_code=400, detail="old_name and new_name are required")
    if old_name not in session.headers:
        raise HTTPException(status_code=400, detail="Column not found")
    if new_name != old_name and new_name in session.headers:
        raise HTTPException(status_code=400, detail="New column name already exists")
    idx = session.headers.index(old_name)
    session.headers[idx] = new_name
    for row in session.cleaned_rows:
        row[new_name] = row.pop(old_name, "")
    return {"status": "ok", "headers": session.headers}


@router.post("/ui/api/columns/delete")
async def delete_column(payload: ColumnDeletePayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    name = payload.column_name.strip()
    if name not in session.headers:
        raise HTTPException(status_code=400, detail="Column not found")
    if len(session.headers) <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last column")
    session.headers = [h for h in session.headers if h != name]
    for row in session.cleaned_rows:
        row.pop(name, None)
    return {"status": "ok", "headers": session.headers}


@router.post("/ui/api/configure")
async def configure(payload: ConfigurePayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if payload.selected_column not in session.headers:
        raise HTTPException(status_code=400, detail="Selected column not found")
    attrs = [a.strip() for a in payload.base_attributes if a.strip()]
    items = [
        str(row.get(payload.selected_column, "")).strip()
        for idx, row in enumerate(session.cleaned_rows)
        if idx not in session.excluded_row_indices
    ]
    items = [item for item in items if item]
    if not items:
        raise HTTPException(status_code=400, detail="Selected column has no values")

    session.selected_column = payload.selected_column
    session.base_attributes = attrs
    session.items = items
    return {"selected_column": payload.selected_column, "base_attributes": attrs, "items_count": len(items)}


@router.post("/ui/api/clusterize/start")
async def start_clusterize(
    payload: ClusterizeStartPayload,
    request: Request,
    background_tasks: BackgroundTasks,
) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if not session.items:
        raise HTTPException(status_code=400, detail="Upload file and save config first")
    session.embedding_provider = payload.embedding_provider.strip().lower()
    session.cluster_profile_provider = _validate_llm_provider(payload.profile_provider)
    try:
        embedding = EmbeddingProvider(session.embedding_provider)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    body = ClusterizeRequest(
        items=session.items,
        base_attributes=session.base_attributes,
        embedding_provider=embedding,
        cluster_profile_provider=ClusterProfileProvider(session.cluster_profile_provider),
    )
    created = await dispatch_clusterize(request, body, background_tasks)
    session.clusterize_task_id = created.task_id
    return {
        "task_id": created.task_id,
        "status": created.status,
        "cluster_profile_provider": session.cluster_profile_provider,
        "embedding_provider": session.embedding_provider,
    }


@router.post("/ui/api/normalize/start")
async def start_normalize(
    payload: StartTaskPayload,
    request: Request,
    background_tasks: BackgroundTasks,
) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if not session.approved_clusters:
        raise HTTPException(status_code=400, detail="Approve clusters first")
    if payload.provider:
        session.normalize_provider = _validate_llm_provider(payload.provider)
    mode = _normalize_mode(payload.mode)
    cluster_attribute_mode = _normalize_cluster_attribute_mode(payload.cluster_attribute_mode)
    all_clusters = _build_normalize_clusters_payload(session)
    selected_indexes = _normalize_selected_cluster_indexes(payload.cluster_indexes, len(all_clusters))
    clusters = [all_clusters[index] for index in selected_indexes]
    if not clusters:
        raise HTTPException(status_code=400, detail="No clusters selected for normalization")

    if cluster_attribute_mode in {"all", "missing"}:
        rewritten: list[ClusterInput] = []
        for cluster in clusters:
            all_items = list(cluster.items)
            if cluster_attribute_mode == "all":
                item_sources = {item_name: "ai" for item_name in all_items}
                item_values = {item_name: {} for item_name in all_items}
            else:
                item_sources = {item_name: "memory" for item_name in all_items}
                item_values = {
                    item_name: dict(cluster.item_values.get(item_name, {}))
                    for item_name in all_items
                }
            rewritten.append(
                ClusterInput(
                    name=cluster.name,
                    attributes=list(cluster.attributes),
                    items=all_items,
                    enriched_name_template=cluster.enriched_name_template,
                    item_sources=item_sources,
                    item_values=item_values,
                    attribute_merge=dict(cluster.attribute_merge or {}),
                    attribute_merge_separators=dict(cluster.attribute_merge_separators or {}),
                )
            )
        clusters = rewritten
    resume_completed_cluster_indexes: list[int] = []
    resume_seed_normalized: list[dict[str, Any]] = []
    resume_seed_clusters_collapsed: list[dict[str, Any]] = []
    resume_expected_count: int | None = None

    if mode == "restart":
        session.normalize_result = None
    elif mode == "resume":
        previous_task_id = session.normalize_task_id
        if not previous_task_id:
            raise HTTPException(status_code=400, detail="No previous normalize task to resume")
        previous_status = await fetch_task_status(previous_task_id, request)
        previous_result = dict(previous_status.result or {})
        if not previous_result:
            raise HTTPException(status_code=400, detail="No partial result to resume from")
        resume_completed_cluster_indexes = _extract_resume_indexes(previous_result, len(clusters))
        if not resume_completed_cluster_indexes:
            raise HTTPException(
                status_code=400,
                detail="Partial result has no completed cluster checkpoints",
            )
        resume_seed_normalized = [
            dict(item)
            for item in list(previous_result.get("normalized") or [])
            if isinstance(item, dict)
        ]
        resume_seed_clusters_collapsed = [
            dict(cluster_payload)
            for cluster_payload in list(previous_result.get("clusters_collapsed") or [])
            if isinstance(cluster_payload, dict)
        ]
        raw_expected = previous_result.get("expected_count")
        if isinstance(raw_expected, int):
            resume_expected_count = raw_expected

    body = NormalizeRequest(
        clusters=clusters,
        llm_provider=NormalizeProvider(session.normalize_provider),
        resume_completed_cluster_indexes=resume_completed_cluster_indexes,
        resume_seed_normalized=resume_seed_normalized,
        resume_seed_clusters_collapsed=resume_seed_clusters_collapsed,
        resume_expected_count=resume_expected_count,
    )
    created = await dispatch_normalize(request, body, background_tasks)
    session.normalize_task_id = created.task_id
    return {
        "task_id": created.task_id,
        "status": created.status,
        "mode": mode,
        "resumed_clusters": len(resume_completed_cluster_indexes),
        "selected_clusters": len(clusters),
        "cluster_attribute_mode": cluster_attribute_mode,
    }


@router.get("/ui/api/task/{session_id}/{task_type}")
async def get_task_status(session_id: str, task_type: str, request: Request) -> dict[str, Any]:
    session = _get_session(session_id)
    task_id = session.clusterize_task_id if task_type == "clusterize" else session.normalize_task_id
    if task_type not in {"clusterize", "normalize"}:
        raise HTTPException(status_code=400, detail="task_type must be clusterize|normalize")
    if not task_id:
        raise HTTPException(status_code=400, detail="Task not started")

    status = await fetch_task_status(task_id, request)
    payload = status.model_dump()
    if task_type == "normalize":
        partial_result = dict(status.result or {})
        if partial_result and bool(partial_result.get("is_partial")):
            session.normalize_result = partial_result
            _apply_collapsed_clusters(session)
    if status.status != "COMPLETED":
        return payload
    if task_type == "clusterize":
        session.clusterize_result = status.result or {}
        if not session.approved_clusters:
            session.approved_clusters = _build_default_clusters(session.clusterize_result)
    else:
        session.normalize_result = status.result or {}
        _apply_collapsed_clusters(session)
    return payload


@router.get("/ui/api/clusters/{session_id}")
async def get_clusters(session_id: str) -> dict[str, Any]:
    return {"clusters": _clusters_for_ui(_get_session(session_id))}


@router.get("/ui/api/memory/clusters")
async def list_memory_clusters(request: Request) -> dict[str, Any]:
    clusters = await asyncio.to_thread(request.app.state.vector_db.list_memory_clusters)
    return {"clusters": clusters}


@router.post("/ui/api/memory/cluster/delete")
async def delete_memory_cluster(payload: MemoryClusterDeletePayload, request: Request) -> dict[str, Any]:
    cluster_name = str(payload.cluster_name or "").strip()
    if not cluster_name:
        raise HTTPException(status_code=400, detail="cluster_name is required")
    deleted = await asyncio.to_thread(request.app.state.vector_db.delete_cluster_items, cluster_name)
    if deleted <= 0:
        raise HTTPException(status_code=404, detail="Memory cluster not found")
    return {"deleted_count": deleted, "cluster_name": cluster_name}


@router.post("/ui/api/memory/item/delete")
async def delete_memory_item(payload: MemoryItemDeletePayload, request: Request) -> dict[str, Any]:
    text = str(payload.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="text is required")
    deleted = await asyncio.to_thread(request.app.state.vector_db.delete_item_by_text, text)
    if deleted <= 0:
        raise HTTPException(status_code=404, detail="Memory item not found")
    return {"deleted_count": deleted, "text": text}


@router.post("/ui/api/memory/search")
async def search_memory(payload: MemorySearchPayload, request: Request) -> dict[str, Any]:
    query = str(payload.query or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    query_vectors = await asyncio.to_thread(request.app.state.vectorizer.get_embeddings, [query])
    if not query_vectors:
        return {"items": []}
    matches = await asyncio.to_thread(
        request.app.state.vector_db.search_local,
        query_vectors[0],
        payload.limit,
        0.0,
    )
    return {"items": matches}


@router.post("/ui/api/clusters/{session_id}/memory/load")
async def load_memory_cluster(
    session_id: str,
    payload: MemoryClusterLoadPayload,
    request: Request,
) -> dict[str, Any]:
    session = _get_session(session_id)
    cluster_name = str(payload.cluster_name or "").strip()
    if not cluster_name:
        raise HTTPException(status_code=400, detail="cluster_name is required")
    points = await asyncio.to_thread(request.app.state.vector_db.load_cluster_items, cluster_name)
    if not points:
        raise HTTPException(status_code=404, detail="Memory cluster not found")
    cluster = _cluster_from_memory_points(cluster_name, points, session.base_attributes)

    existing_idx = next(
        (
            idx
            for idx, current in enumerate(session.approved_clusters or [])
            if str(current.get("memory_cluster_name") or "").strip() == cluster_name
        ),
        None,
    )
    if existing_idx is None:
        session.approved_clusters.append(cluster)
        existing_idx = len(session.approved_clusters) - 1
    else:
        session.approved_clusters[existing_idx] = cluster
    return {"cluster": _clusters_for_ui(session)[existing_idx], "cluster_index": existing_idx}


@router.post("/ui/api/clusters/{session_id}/{cluster_index}/memory/load-full")
async def load_cluster_full_from_memory(
    session_id: str,
    cluster_index: int,
    request: Request,
) -> dict[str, Any]:
    session = _get_session(session_id)
    if cluster_index < 0 or cluster_index >= len(session.approved_clusters):
        raise HTTPException(status_code=400, detail="Cluster index out of range")
    cluster = session.approved_clusters[cluster_index]
    cluster_name = str(cluster.get("memory_cluster_name") or cluster.get("name") or "").strip()
    if not cluster_name:
        raise HTTPException(status_code=400, detail="Memory cluster name is empty")
    points = await asyncio.to_thread(request.app.state.vector_db.load_cluster_items, cluster_name)
    if not points:
        raise HTTPException(status_code=404, detail="Memory cluster not found")
    session.approved_clusters[cluster_index] = _cluster_from_memory_points(
        cluster_name,
        points,
        session.base_attributes,
    )
    return {"cluster": _clusters_for_ui(session)[cluster_index], "cluster_index": cluster_index}


@router.post("/ui/api/clusters/save")
async def save_clusters(payload: ClustersPayload) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    valid: list[dict[str, Any]] = []
    memory_items = _memory_item_names(session)
    for cluster in payload.clusters:
        name = str(cluster.get("name", "")).strip() or "Cluster"
        source = str(cluster.get("source") or "").strip().lower()
        if source not in ("memory", "ai", "manual"):
            source = "ai"
        memory_cluster_name = str(cluster.get("memory_cluster_name") or "").strip()
        attrs = [str(v).strip() for v in cluster.get("attributes", []) if str(v).strip()]
        items = [str(v).strip() for v in cluster.get("items", []) if str(v).strip()]
        if not items:
            continue
        payload_rows = cluster.get("rows") or []
        rows: list[dict[str, Any]] = []
        if payload_rows:
            for row in payload_rows:
                item_name = str(row.get("item", "")).strip()
                if not item_name:
                    continue
                source = str(row.get("source") or "").strip().lower()
                if source not in ("memory", "ai"):
                    source = "memory" if item_name in memory_items else "ai"
                values = row.get("values")
                row_values = values if isinstance(values, dict) else {}
                enriched = str(row.get("enriched_name") or "").strip()
                aliases_raw = row.get("aliases") or []
                aliases = [
                    str(alias).strip() for alias in aliases_raw if str(alias).strip()
                ]
                row_payload: dict[str, Any] = {
                    "item": item_name,
                    "values": {
                        str(k).strip(): str(v).strip()
                        for k, v in row_values.items()
                        if str(k).strip()
                    },
                    "source": source,
                }
                if enriched:
                    row_payload["enriched_name"] = enriched
                if aliases:
                    row_payload["aliases"] = aliases
                members_raw = row.get("members") or []
                members = [
                    {
                        "item": str(member.get("item", "")).strip(),
                        "values": {
                            str(k).strip(): str(v).strip()
                            for k, v in dict(member.get("values") or {}).items()
                            if str(k).strip()
                        },
                        "source": str(member.get("source") or source).strip().lower()
                        if str(member.get("source") or source).strip().lower()
                        in ("memory", "ai")
                        else source,
                    }
                    for member in members_raw
                    if str(member.get("item", "")).strip()
                ]
                if members:
                    row_payload["members"] = members
                rows.append(row_payload)
        else:
            rows = [
                {
                    "item": item_name,
                    "values": {},
                    "source": "memory" if item_name in memory_items else "ai",
                }
                for item_name in items
            ]
        attribute_merge = _normalize_attribute_merge(cluster.get("attribute_merge"))
        attribute_merge_separators = _normalize_attribute_merge_separators(
            cluster.get("attribute_merge_separators")
        )
        valid.append(
            {
                "name": name,
                "attributes": attrs or session.base_attributes,
                "enriched_name_template": str(
                    cluster.get("enriched_name_template") or ""
                ).strip(),
                "attribute_merge": attribute_merge,
                "attribute_merge_separators": attribute_merge_separators,
                "items": items,
                "rows": rows,
                "source": _cluster_source({"rows": rows, "source": source}),
                "memory_cluster_name": memory_cluster_name,
            }
        )
    if not valid:
        raise HTTPException(status_code=400, detail="No valid clusters to save")
    session.approved_clusters = valid
    return {"saved_clusters": len(valid)}


@router.post("/ui/api/clusters/{session_id}/rededupe")
async def rededupe_clusters(session_id: str, request: Request) -> dict[str, Any]:
    """Пересчитывает дубликаты обогащённых имён в утверждённых кластерах."""
    session = _get_session(session_id)
    if not session.approved_clusters:
        raise HTTPException(status_code=400, detail="No clusters to rededupe")

    provider_name = str(session.embedding_provider or EmbeddingProvider.LOCAL.value)
    try:
        provider = EmbeddingProvider(provider_name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    _provider_label, embedding_client = _get_embedding_client(request, provider)

    updated_clusters: list[dict[str, Any]] = []
    for cluster in session.approved_clusters:
        template = str(cluster.get("enriched_name_template") or "").strip()
        attribute_merge = _normalize_attribute_merge(cluster.get("attribute_merge"))
        attribute_merge_separators = _normalize_attribute_merge_separators(
            cluster.get("attribute_merge_separators")
        )
        entries: list[dict[str, Any]] = []
        for row in cluster.get("rows") or []:
            entries.extend(_row_entries_for_rededupe(row, template))
        if not entries:
            updated_clusters.append(cluster)
            continue
        texts = [
            str(entry.get("enriched_name") or entry.get("item") or "").strip()
            for entry in entries
        ]
        embeddings = await asyncio.to_thread(embedding_client.get_embeddings, texts)
        collapsed = await asyncio.to_thread(
            collapse_cluster_rows,
            entries,
            embeddings,
            None,
            attribute_merge,
            attribute_merge_separators,
        )
        merged = dict(cluster)
        merged["rows"] = collapsed
        merged["items"] = [
            str(row.get("enriched_name") or row.get("item") or "").strip()
            for row in collapsed
            if str(row.get("enriched_name") or row.get("item") or "").strip()
        ]
        updated_clusters.append(merged)

    session.approved_clusters = updated_clusters
    return {"clusters": _clusters_for_ui(session)}


@router.post("/ui/api/memory/save")
async def save_memory(payload: SaveMemoryPayload, request: Request) -> dict[str, Any]:
    session = _get_session(payload.session_id)
    if not session.approved_clusters:
        raise HTTPException(status_code=400, detail="No clusters to map items to memory")
    selected_cluster_index = payload.cluster_index
    if selected_cluster_index is not None:
        if selected_cluster_index < 0 or selected_cluster_index >= len(session.approved_clusters):
            raise HTTPException(status_code=400, detail="Cluster index out of range")
        clusters_for_save = [session.approved_clusters[selected_cluster_index]]
    else:
        clusters_for_save = list(session.approved_clusters)
    normalized = (session.normalize_result or {}).get("normalized") or []
    cluster_by_item = _item_cluster_name_map(clusters_for_save)
    items: list[MemoryItem] = []
    for cluster in clusters_for_save:
        cluster_name = (
            str(cluster.get("memory_cluster_name") or "").strip()
            or str(cluster.get("name", "")).strip()
            or "Cluster"
        )
        cluster_attribute_merge = _normalize_attribute_merge(cluster.get("attribute_merge"))
        cluster_attribute_merge_separators = _normalize_attribute_merge_separators(
            cluster.get("attribute_merge_separators")
        )
        for row in cluster.get("rows") or []:
            enriched = str(row.get("enriched_name") or "").strip()
            if not enriched:
                continue
            aliases = [
                str(alias).strip()
                for alias in row.get("aliases") or []
                if str(alias).strip()
            ]
            original_item_values: dict[str, dict[str, Any]] = {}
            members = row.get("members") or []
            for member in members:
                item_name = str(member.get("item", "")).strip()
                if not item_name:
                    continue
                member_values = {
                    str(key).strip(): str(value).strip()
                    for key, value in dict(member.get("values") or {}).items()
                    if str(key).strip()
                }
                original_item_values[item_name] = member_values
            for alias in aliases:
                if alias not in original_item_values:
                    original_item_values[alias] = {
                        str(key).strip(): str(value).strip()
                        for key, value in dict(row.get("values") or {}).items()
                        if str(key).strip()
                    }
            items.append(
                MemoryItem(
                    text=enriched,
                    attributes=dict(row.get("values") or {}),
                    cluster_name=cluster_name,
                    original_items=aliases or [enriched],
                    original_item_values=original_item_values,
                    attribute_merge=cluster_attribute_merge,
                    attribute_merge_separators=cluster_attribute_merge_separators,
                )
            )
    if not items and normalized:
        for row in normalized:
            item_name = str(row.get("item", "")).strip()
            if not item_name:
                continue
            cluster_name = cluster_by_item.get(item_name)
            if not cluster_name:
                raise HTTPException(
                    status_code=400,
                    detail=f"Item not found in approved clusters: {item_name}",
                )
            items.append(
                MemoryItem(
                    text=str(row.get("enriched_name") or item_name).strip(),
                    attributes=dict(row.get("values", {})),
                    cluster_name=cluster_name,
                    original_items=[item_name],
                    original_item_values={item_name: dict(row.get("values", {}))},
                    attribute_merge={},
                    attribute_merge_separators={},
                )
            )
    if not items:
        raise HTTPException(status_code=400, detail="No valid normalized items")
    if selected_cluster_index is not None:
        cluster_name = (
            str(clusters_for_save[0].get("memory_cluster_name") or "").strip()
            or str(clusters_for_save[0].get("name", "")).strip()
            or "Cluster"
        )
        await asyncio.to_thread(request.app.state.vector_db.delete_cluster_items, cluster_name)
    saved = await dispatch_memory_save(
        request,
        MemorySaveRequest(items=items),
        embedding_provider=session.embedding_provider,
    )
    return {"saved_count": saved.saved_count}


@router.get("/ui/api/export/{session_id}/xlsx")
async def export_xlsx(session_id: str) -> StreamingResponse:
    session = _get_session(session_id)
    clusters = session.approved_clusters or []
    if not clusters:
        raise HTTPException(status_code=400, detail="No clusters to export")
    normalized_map = _normalized_map(session)

    wb = Workbook()
    wb.remove(wb.active)
    for idx, cluster in enumerate(clusters, start=1):
        sheet_name = (str(cluster.get("name", "")).strip() or f"Cluster_{idx}")[:31]
        ws = wb.create_sheet(title=sheet_name)
        attributes, rows_out = _cluster_export_rows(cluster, normalized_map)
        ws.append(["Обогащенное наименование", "Исходные номенклатуры", *attributes])
        for export_row in rows_out:
            ws.append(export_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="clusters_export.xlsx"'},
    )


@router.get("/ui/api/export/{session_id}/cluster/{cluster_index}/xlsx")
async def export_cluster_xlsx(session_id: str, cluster_index: int) -> StreamingResponse:
    session = _get_session(session_id)
    cluster = _get_cluster_for_export(session, cluster_index)
    normalized_map = _normalized_map(session)

    wb = Workbook()
    ws = wb.active
    ws.title = (str(cluster.get("name", "")).strip() or f"Cluster_{cluster_index + 1}")[:31]
    attributes, rows_out = _cluster_export_rows(cluster, normalized_map)
    ws.append(["Обогащенное наименование", "Исходные номенклатуры", *attributes])
    for export_row in rows_out:
        ws.append(export_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    cluster_name = str(cluster.get("name", "")).strip() or f"cluster_{cluster_index + 1}"
    filename = f"{cluster_name}.xlsx".replace('"', "")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ui/api/export/{session_id}/cluster/{cluster_index}/csv")
async def export_cluster_csv(session_id: str, cluster_index: int) -> StreamingResponse:
    session = _get_session(session_id)
    cluster = _get_cluster_for_export(session, cluster_index)
    normalized_map = _normalized_map(session)
    attributes, rows_out = _cluster_export_rows(cluster, normalized_map)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Обогащенное наименование", "Исходные номенклатуры", *attributes])
    for export_row in rows_out:
        writer.writerow(export_row)
    csv_text = output.getvalue()
    output.close()
    data = io.BytesIO(csv_text.encode("utf-8-sig"))

    cluster_name = str(cluster.get("name", "")).strip() or f"cluster_{cluster_index + 1}"
    filename = f"{cluster_name}.csv".replace('"', "")
    return StreamingResponse(
        data,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/ui/api/export/{session_id}/csv")
async def export_csv_clusters_zip(session_id: str) -> StreamingResponse:
    session = _get_session(session_id)
    clusters = session.approved_clusters or []
    if not clusters:
        raise HTTPException(status_code=400, detail="No clusters to export")
    normalized_map = _normalized_map(session)

    archive_stream = io.BytesIO()
    with zipfile.ZipFile(archive_stream, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for idx, cluster in enumerate(clusters, start=1):
            attributes, rows_out = _cluster_export_rows(cluster, normalized_map)
            csv_stream = io.StringIO()
            writer = csv.writer(csv_stream)
            writer.writerow(["Обогащенное наименование", "Исходные номенклатуры", *attributes])
            for export_row in rows_out:
                writer.writerow(export_row)
            cluster_name = str(cluster.get("name", "")).strip()
            safe_name = _safe_filename(cluster_name, f"cluster_{idx}")
            archive.writestr(f"{idx:02d}_{safe_name}.csv", csv_stream.getvalue().encode("utf-8-sig"))

    archive_stream.seek(0)
    return StreamingResponse(
        archive_stream,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="clusters_csv_export.zip"'},
    )


@router.post("/ui/api/admin/flush-redis")
async def flush_redis_cache(request: Request) -> dict[str, Any]:
    """Очищает Redis: статусы задач и кэш ответов LLM."""
    await request.app.state.redis.flushdb()
    request.app.state.llm_clients = {}
    return {
        "status": "ok",
        "message": "Redis очищен (задачи и кэш LLM)",
    }


@router.post("/ui/api/admin/flush-qdrant")
async def flush_qdrant_memory(request: Request) -> dict[str, Any]:
    """Удаляет коллекцию векторной памяти в Qdrant."""
    vector_db = request.app.state.vector_db
    points_before = vector_db.get_points_count()
    collection_name = vector_db.collection_name
    removed = vector_db.clear_collection()
    if removed:
        message = f"Коллекция «{collection_name}» удалена ({points_before} точек)"
    else:
        message = f"Коллекция «{collection_name}» не найдена — уже пусто"
    return {
        "status": "ok",
        "message": message,
        "collection": collection_name,
        "points_before": points_before,
        "removed": removed,
    }
